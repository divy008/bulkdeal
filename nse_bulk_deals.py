import os
import requests
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIG - set these as GitHub Actions repo secrets, read via env vars
# ---------------------------------------------------------------------------
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "")
EXCEL_PATH = "bulk_deals_report.xlsx"

# ---------------------------------------------------------------------------
# 1. FETCH — today only, no date range
# ---------------------------------------------------------------------------
def fetch_nse_data(date_str):
    """date_str format: DD-MM-YYYY. Always just one day."""
    base_url = "https://www.nseindia.com"
    headers = {'User-Agent': 'Mozilla/5.0', 'Referer': f'{base_url}/report-detail/display-bulk-and-block-deals'}

    session = requests.Session()
    session.get(base_url, headers=headers)

    api_url = f"{base_url}/api/historicalOR/bulk-block-short-deals?optionType=bulk_deals&from={date_str}&to={date_str}"
    try:
        response = session.get(api_url, headers=headers)
        if response.status_code == 200:
            return response.json().get('data', [])
    except Exception:
        pass
    return []


# ---------------------------------------------------------------------------
# 2. PROCESS (same logic as before, kept intact)
# ---------------------------------------------------------------------------
def process_data(data):
    negotiated_indices = set()
    transfers = []
    for i in range(len(data)):
        if i in negotiated_indices:
            continue
        for j in range(i + 1, len(data)):
            if j in negotiated_indices:
                continue
            if (data[i]['BD_SYMBOL'] == data[j]['BD_SYMBOL'] and
                data[i]['BD_QTY_TRD'] == data[j]['BD_QTY_TRD'] and
                data[i]['BD_DT_DATE'] == data[j]['BD_DT_DATE'] and
                data[i]['BD_BUY_SELL'] != data[j]['BD_BUY_SELL']):
                transfers.append((data[i], data[j]))
                negotiated_indices.update([i, j])
                break

    negotiated_symbols = {data[i]['BD_SYMBOL'] for i in negotiated_indices}
    strategic_data = [data[i] for i in range(len(data)) if i not in negotiated_indices]

    tracker = {}
    for d in strategic_data:
        key = (d['BD_SYMBOL'], d['BD_CLIENT_NAME'])
        if key not in tracker:
            tracker[key] = {'BUY': 0, 'SELL': 0, 'records': []}
        tracker[key][d['BD_BUY_SELL']] += d['BD_QTY_TRD']
        tracker[key]['records'].append(d)

    intraday, directional = [], []
    for key, info in tracker.items():
        if info['BUY'] > 0 and info['SELL'] > 0:
            intraday.extend(info['records'])
        else:
            directional.extend(info['records'])
    directional = [d for d in directional if d['BD_SYMBOL'] not in negotiated_symbols]

    volume_comp = {}
    client_count_per_symbol = {}
    for d in data:
        sym = d['BD_SYMBOL']
        if sym not in volume_comp:
            volume_comp[sym] = {'BUY_QTY': 0, 'SELL_QTY': 0, 'BUY_VAL': 0, 'SELL_VAL': 0}
            client_count_per_symbol[sym] = {'BUY': set(), 'SELL': set()}
        val = d['BD_QTY_TRD'] * d['BD_TP_WATP']
        client_count_per_symbol[sym][d['BD_BUY_SELL']].add(d['BD_CLIENT_NAME'])
        if d['BD_BUY_SELL'] == 'BUY':
            volume_comp[sym]['BUY_QTY'] += d['BD_QTY_TRD']
            volume_comp[sym]['BUY_VAL'] += val
        else:
            volume_comp[sym]['SELL_QTY'] += d['BD_QTY_TRD']
            volume_comp[sym]['SELL_VAL'] += val

    return {
        "transfers": transfers,
        "intraday": intraday,
        "directional": directional,
        "volume_comp": volume_comp,
        "client_count_per_symbol": client_count_per_symbol,
    }


# ---------------------------------------------------------------------------
# 3. OPPORTUNITY RANKING - the "who's being accumulated today" signal
# ---------------------------------------------------------------------------
def rank_opportunities(volume_comp, client_count_per_symbol, top_n=5):
    rows = []
    for sym, v in volume_comp.items():
        net_qty = v['BUY_QTY'] - v['SELL_QTY']
        avg_buy = v['BUY_VAL'] / v['BUY_QTY'] if v['BUY_QTY'] else 0
        avg_sell = v['SELL_VAL'] / v['SELL_QTY'] if v['SELL_QTY'] else 0
        buyers = len(client_count_per_symbol[sym]['BUY'])
        sellers = len(client_count_per_symbol[sym]['SELL'])
        rows.append({
            "symbol": sym, "net_qty": net_qty, "avg_buy": avg_buy, "avg_sell": avg_sell,
            "buyers": buyers, "sellers": sellers,
        })

    accumulation = sorted([r for r in rows if r['net_qty'] > 0], key=lambda r: r['net_qty'], reverse=True)
    distribution = sorted([r for r in rows if r['net_qty'] < 0], key=lambda r: r['net_qty'])

    # Multi-buyer conviction: more than one distinct client net-buying same stock same day
    multi_buyer = [r for r in accumulation if r['buyers'] > 1]

    return accumulation[:top_n], distribution[:top_n], multi_buyer[:top_n]


# ---------------------------------------------------------------------------
# 4. TELEGRAM MESSAGE (short summary, HTML formatted)
# ---------------------------------------------------------------------------
def build_telegram_message(date_str, accumulation, distribution, multi_buyer, transfers):
    lines = [f"<b>📊 NSE Bulk Deals — {date_str}</b>\n"]

    if multi_buyer:
        lines.append("<b>🔥 Multiple Buyers Piling In (strongest signal)</b>")
        for r in multi_buyer:
            lines.append(f"• <b>{r['symbol']}</b> — {r['buyers']} buyers, net +{r['net_qty']:,} @ ₹{r['avg_buy']:.2f}")
        lines.append("")

    if accumulation:
        lines.append("<b>🟢 Top Accumulation (Net Buy)</b>")
        for r in accumulation:
            lines.append(f"• {r['symbol']}: +{r['net_qty']:,} @ ₹{r['avg_buy']:.2f}")
        lines.append("")

    if distribution:
        lines.append("<b>🔴 Top Distribution (Net Sell)</b>")
        for r in distribution:
            lines.append(f"• {r['symbol']}: {r['net_qty']:,} @ ₹{r['avg_sell']:.2f}")
        lines.append("")

    if transfers:
        lines.append(f"<b>🔁 Negotiated Transfers:</b> {len(transfers)} block(s) — see Excel for detail")

    lines.append("\n<i>Full breakdown attached as Excel.</i>")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# 5. EXCEL EXPORT (full detail, one sheet per table)
# ---------------------------------------------------------------------------
def export_excel(path, intraday, directional, volume_comp, client_count_per_symbol, transfers):
    def style_header(ws):
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="305496")
        for i, col in enumerate(ws.columns, 1):
            width = max(len(str(c.value)) for c in col if c.value is not None) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(width, 40)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(intraday).to_excel(writer, sheet_name="Intraday_Arbitrage", index=False)

        pd.DataFrame(directional).to_excel(writer, sheet_name="Strategic_Activity", index=False)

        opp_rows = []
        for sym, v in volume_comp.items():
            net = v['BUY_QTY'] - v['SELL_QTY']
            avg_b = v['BUY_VAL'] / v['BUY_QTY'] if v['BUY_QTY'] else 0
            avg_s = v['SELL_VAL'] / v['SELL_QTY'] if v['SELL_QTY'] else 0
            opp_rows.append({
                "Symbol": sym, "Net Qty": net, "Buy Qty": v['BUY_QTY'], "Sell Qty": v['SELL_QTY'],
                "Avg Buy Price": round(avg_b, 2), "Avg Sell Price": round(avg_s, 2),
                "Distinct Buyers": len(client_count_per_symbol[sym]['BUY']),
                "Distinct Sellers": len(client_count_per_symbol[sym]['SELL']),
            })
        opp_df = pd.DataFrame(opp_rows).sort_values("Net Qty", ascending=False)
        opp_df.to_excel(writer, sheet_name="Opportunities", index=False)

        transfer_rows = []
        for d1, d2 in transfers:
            s, b = (d1, d2) if d1['BD_BUY_SELL'] == 'SELL' else (d2, d1)
            transfer_rows.append({
                "Date": d1['BD_DT_DATE'], "Symbol": d1['BD_SYMBOL'], "Qty": d1['BD_QTY_TRD'],
                "Seller": s['BD_CLIENT_NAME'], "Buyer": b['BD_CLIENT_NAME'],
            })
        pd.DataFrame(transfer_rows).to_excel(writer, sheet_name="Negotiated_Transfers", index=False)

        for sheet in writer.sheets.values():
            if sheet.max_row >= 1 and sheet.max_column >= 1:
                style_header(sheet)


# ---------------------------------------------------------------------------
# 6. TELEGRAM SEND
# ---------------------------------------------------------------------------
def send_telegram_message(text):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("Telegram credentials missing, skipping send.")
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    requests.post(url, data={"chat_id": TELEGRAM_CHAT_ID, "text": text, "parse_mode": "HTML"})


def send_telegram_document(path, caption=""):
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        print("Telegram credentials missing, skipping send.")
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    with open(path, "rb") as f:
        requests.post(url, data={"chat_id": TELEGRAM_CHAT_ID, "caption": caption}, files={"document": f})


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    today = datetime.now().strftime("%d-%m-%Y")
    data = fetch_nse_data(today)
    if not data:
        send_telegram_message(f"📊 NSE Bulk Deals — {today}\nNo bulk deals reported today.")
        return

    result = process_data(data)
    accumulation, distribution, multi_buyer = rank_opportunities(
        result["volume_comp"], result["client_count_per_symbol"]
    )

    export_excel(
        EXCEL_PATH, result["intraday"], result["directional"], result["volume_comp"],
        result["client_count_per_symbol"], result["transfers"]
    )

    msg = build_telegram_message(today, accumulation, distribution, multi_buyer, result["transfers"])
    send_telegram_message(msg)
    send_telegram_document(EXCEL_PATH, caption=f"Full bulk deals detail — {today}")


if __name__ == "__main__":
    main()
